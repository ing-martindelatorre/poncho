#!/usr/bin/env python3
"""Import OMAR SEMANA 27.xlsx data into the Poncho database."""

import re
import uuid
from datetime import datetime, date

import openpyxl
import psycopg2

XLSX_PATH = "/home/zidane/xyron/proyectos/poncho/OMAR SEMANA 27.xlsx"
DB_URL = "postgresql://poncho:poncho_secure_2026_xK9m@localhost:5434/poncho"

MONTHS_ES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def cuid():
    return "cx" + uuid.uuid4().hex[:23]


def parse_date_range(text):
    """Parse '08-13 DE DICIEMBRE DEL 2025' into (start_date, end_date)."""
    text = text.strip().upper()
    m = re.match(r"(\d{1,2})-(\d{1,2})\s+DE\s+(\w+)\s+DEL?\s+(\d{4})", text)
    if not m:
        return None, None
    day_start = int(m.group(1))
    day_end = int(m.group(2))
    month_name = m.group(3)
    year = int(m.group(4))
    month = MONTHS_ES.get(month_name)
    if not month:
        return None, None
    start = date(year, month, day_start)
    # Handle month rollover (e.g., "30-04 DE ABRIL" means Mar30-Apr4)
    try:
        end = date(year, month, day_end)
    except ValueError:
        if day_end < day_start:
            next_month = month + 1 if month < 12 else 1
            next_year = year if month < 12 else year + 1
            end = date(next_year, next_month, day_end)
        else:
            end = date(year, month, min(day_end, 28))
    if end < start:
        next_month = month + 1 if month < 12 else 1
        next_year = year if month < 12 else year + 1
        end = date(next_year, next_month, day_end)
    return start, end


def num(val):
    """Convert cell value to float, default 0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    now = datetime.now()

    # --- 1. Create Project ---
    total_cost = 3566708.98
    cost_per_m2 = 8124.62
    built_m2 = round(total_cost / cost_per_m2, 2)

    project_id = cuid()
    cur.execute("""
        INSERT INTO projects (id, name, address, client_name, built_area_m2, status, start_date, notes, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, 'ACTIVE', %s, %s, %s, %s)
    """, (
        project_id,
        "Real Alcazar 1370",
        "Fracc. Jardines Providencia, Tepic",
        "Omar",
        built_m2,
        date(2025, 12, 8),
        "Importado desde OMAR SEMANA 27.xlsx",
        now, now,
    ))
    print(f"Proyecto creado: Real Alcazar 1370 ({built_m2} m2)")

    # --- 2. Parse DESGLOZADO (27 weeks breakdown) ---
    ws = wb["DESGLOZADO"]
    weeks_data = []
    for r in range(14, 41):
        week_num = ws.cell(r, 1).value
        week_label = ws.cell(r, 2).value
        if week_num is None or week_label is None:
            continue
        week_num = int(week_num)
        start_date, end_date = parse_date_range(str(week_label))
        if not start_date:
            continue

        weeks_data.append({
            "num": week_num,
            "label": str(week_label).strip(),
            "start": start_date,
            "end": end_date,
            "permisos": num(ws.cell(r, 3).value),
            "mano_obra": num(ws.cell(r, 4).value),
            "material_efe": num(ws.cell(r, 5).value),
            "material_fact": num(ws.cell(r, 6).value),
            "honorarios": num(ws.cell(r, 7).value),
            "soldador": num(ws.cell(r, 8).value),
            "electricista": num(ws.cell(r, 9).value),
            "aa": num(ws.cell(r, 10).value),
            "pintor": num(ws.cell(r, 11).value),
            "nomina_3pct": num(ws.cell(r, 12).value),
            "imss": num(ws.cell(r, 13).value),
            "fierrero": num(ws.cell(r, 14).value),
            "fontanero": num(ws.cell(r, 15).value),
        })

    # --- 3. Parse TOTALES (for payment/abono data) ---
    ws_totales = wb["TOTALES"]
    totals_by_week = {}
    for r in range(12, 39):
        wn = ws_totales.cell(r, 1).value
        if wn is None:
            continue
        totals_by_week[int(wn)] = {
            "efectivo": num(ws_totales.cell(r, 3).value),
            "facturado": num(ws_totales.cell(r, 4).value),
            "honorarios": num(ws_totales.cell(r, 5).value),
            "total": num(ws_totales.cell(r, 6).value),
        }

    # --- 4. Create WeeklyPeriods and entries ---
    period_ids = {}
    for w in weeks_data:
        period_id = cuid()
        period_ids[w["num"]] = period_id
        cur.execute("""
            INSERT INTO weekly_periods (id, project_id, week_number, label, start_date, end_date, status, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, 'CLOSED', %s, %s)
        """, (period_id, project_id, w["num"], w["label"], w["start"], w["end"], now, now))

        # --- Labor Payment (Mano de obra) ---
        if w["mano_obra"] > 0:
            cur.execute("""
                INSERT INTO labor_payments (id, weekly_period_id, worker_name, role, rate, total, money_kind, notes, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, 'CASH', %s, %s, %s)
            """, (cuid(), period_id, "Cuadrilla semanal", "Albañilería", w["mano_obra"], w["mano_obra"],
                  f"Importado semana {w['num']}", now, now))

        # --- Material Efectivo ---
        if w["material_efe"] > 0:
            cur.execute("""
                INSERT INTO material_purchases (id, project_id, weekly_period_id, description, total, paid_amount, money_kind, status, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, 'CASH', 'DELIVERED', %s, %s)
            """, (cuid(), project_id, period_id, f"Material efectivo semana {w['num']}",
                  w["material_efe"], w["material_efe"], now, now))

        # --- Material Facturado ---
        if w["material_fact"] > 0:
            cur.execute("""
                INSERT INTO material_purchases (id, project_id, weekly_period_id, description, total, paid_amount, money_kind, status, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, 'INVOICED', 'DELIVERED', %s, %s)
            """, (cuid(), project_id, period_id, f"Material facturado semana {w['num']}",
                  w["material_fact"], w["material_fact"], now, now))

        # --- Permisos ---
        if w["permisos"] > 0:
            cur.execute("""
                INSERT INTO work_items (id, weekly_period_id, category, description, unit, volume, unit_price, total, money_kind, created_at, updated_at)
                VALUES (%s, %s, 'Permisos', %s, 'LOTE', 1, %s, %s, 'CASH', %s, %s)
            """, (cuid(), period_id, f"Permisos semana {w['num']}", w["permisos"], w["permisos"], now, now))

        # --- Subcontratos ---
        subcontracts = [
            ("soldador", "Soldador"), ("electricista", "Electricista"),
            ("aa", "Aire Acondicionado"), ("pintor", "Pintor"),
            ("fierrero", "Fierrero"), ("fontanero", "Fontanero"),
        ]
        for key, name in subcontracts:
            if w[key] > 0:
                cur.execute("""
                    INSERT INTO work_items (id, weekly_period_id, category, description, unit, volume, unit_price, total, money_kind, created_at, updated_at)
                    VALUES (%s, %s, 'Subcontrato', %s, 'LOTE', 1, %s, %s, 'CASH', %s, %s)
                """, (cuid(), period_id, f"{name} semana {w['num']}", w[key], w[key], now, now))

        # --- 3% Nómina ---
        if w["nomina_3pct"] > 0:
            cur.execute("""
                INSERT INTO work_items (id, weekly_period_id, category, description, unit, volume, unit_price, total, money_kind, created_at, updated_at)
                VALUES (%s, %s, 'Nómina 3%%', %s, 'LOTE', 1, %s, %s, 'CASH', %s, %s)
            """, (cuid(), period_id, f"3% nómina semana {w['num']}", w["nomina_3pct"], w["nomina_3pct"], now, now))

        # --- IMSS ---
        if w["imss"] > 0:
            cur.execute("""
                INSERT INTO work_items (id, weekly_period_id, category, description, unit, volume, unit_price, total, money_kind, created_at, updated_at)
                VALUES (%s, %s, 'IMSS', %s, 'LOTE', 1, %s, %s, 'CASH', %s, %s)
            """, (cuid(), period_id, f"IMSS semana {w['num']}", w["imss"], w["imss"], now, now))

    print(f"Creadas {len(weeks_data)} semanas con desglose")

    # --- 5. Week 27 detailed DESTAJOS ---
    ws_dest = wb["DESTAJOS"]
    period_27 = period_ids.get(27)
    destajos_count = 0
    if period_27:
        # Delete the generic labor entry for week 27 — we have detail
        cur.execute("DELETE FROM labor_payments WHERE weekly_period_id = %s", (period_27,))

        for r in range(12, 28):
            desc = ws_dest.cell(r, 2).value
            unit = ws_dest.cell(r, 3).value
            volume = num(ws_dest.cell(r, 8).value)
            price = num(ws_dest.cell(r, 9).value)
            total = num(ws_dest.cell(r, 10).value)
            if not desc or total == 0:
                continue
            desc = str(desc).strip()
            unit = str(unit).strip() if unit else "LOTE"

            cur.execute("""
                INSERT INTO work_items (id, weekly_period_id, category, description, unit, volume, unit_price, total, money_kind, notes, created_at, updated_at)
                VALUES (%s, %s, 'Destajo', %s, %s, %s, %s, %s, 'CASH', %s, %s, %s)
            """, (cuid(), period_27, desc, unit, volume, price, total,
                  "Detalle semana 27 importado", now, now))
            destajos_count += 1

    print(f"Insertados {destajos_count} destajos detallados (semana 27)")

    # --- 6. Week 27 CARATULA EFE details ---
    if period_27:
        # Material efectivo details
        ws_efe = wb["CARATULA EFE"]
        efe_materials = [
            (ws_efe.cell(26, 2).value, num(ws_efe.cell(26, 3).value)),  # GASOLINA
            (ws_efe.cell(27, 2).value, num(ws_efe.cell(27, 3).value)),  # REGLAS
        ]
        # Delete generic material entry for week 27 — we have detail
        cur.execute("""
            DELETE FROM material_purchases WHERE weekly_period_id = %s AND money_kind = 'CASH'
            AND description LIKE 'Material efectivo semana 27'
        """, (period_27,))

        for desc, amount in efe_materials:
            if desc and amount > 0:
                cur.execute("""
                    INSERT INTO material_purchases (id, project_id, weekly_period_id, description, total, paid_amount, money_kind, status, notes, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, 'CASH', 'DELIVERED', 'Detalle semana 27', %s, %s)
                """, (cuid(), project_id, period_27, str(desc).strip(), amount, amount, now, now))

        # Subcontratos efectivo details
        sub_items = [
            (ws_efe.cell(37, 2).value, num(ws_efe.cell(37, 3).value)),  # ELECTRICISTA
            (ws_efe.cell(38, 2).value, num(ws_efe.cell(38, 3).value)),  # IMSS
        ]
        for desc, amount in sub_items:
            if desc and amount > 0:
                desc_str = str(desc).strip()
                # These are already covered by DESGLOZADO subcontract entries
                # Skip to avoid duplication

        # Material facturado detail from CARATULA FACT
        ws_fact = wb["CARATULA FACT"]
        fact_desc = ws_fact.cell(15, 2).value
        fact_amount = num(ws_fact.cell(15, 3).value)
        if fact_desc and fact_amount > 0:
            # Delete generic entry and add detailed one
            cur.execute("""
                DELETE FROM material_purchases WHERE weekly_period_id = %s AND money_kind = 'INVOICED'
                AND description LIKE 'Material facturado semana 27'
            """, (period_27,))
            cur.execute("""
                INSERT INTO material_purchases (id, project_id, weekly_period_id, description, invoice_number, total, paid_amount, money_kind, status, notes, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 'INVOICED', 'DELIVERED', 'Detalle semana 27', %s, %s)
            """, (cuid(), project_id, period_27, str(fact_desc).strip(), "F.24518",
                  fact_amount, fact_amount, now, now))

    print("Detalles semana 27 (materiales efectivo y facturado)")

    # --- 7. PREPAGOS (supplier tracking) ---
    ws_pre = wb["PREPAGOS"]
    suppliers = {
        "El Gallo": {"col": 3, "prepago": num(ws_pre.cell(12, 3).value)},
        "Baez": {"col": 4, "prepago": num(ws_pre.cell(12, 4).value)},
    }

    for supplier_name, info in suppliers.items():
        # Create supplier
        supplier_id = cuid()
        cur.execute("""
            INSERT INTO suppliers (id, name, notes, active, created_at, updated_at)
            VALUES (%s, %s, %s, true, %s, %s)
        """, (supplier_id, supplier_name, f"Prepago total: ${info['prepago']:,.2f}", now, now))

        # Create prepago material purchase
        purchase_id = cuid()
        total_delivered = 0
        deliveries = []
        for r in range(15, 42):
            wn = ws_pre.cell(r, 1).value
            amount = num(ws_pre.cell(r, info["col"]).value)
            week_label = ws_pre.cell(r, 2).value
            if wn is not None and amount > 0:
                start_d, end_d = parse_date_range(str(week_label))
                if start_d:
                    deliveries.append((int(wn), amount, start_d))
                    total_delivered += amount

        if info["prepago"] > 0:
            cur.execute("""
                INSERT INTO material_purchases (id, project_id, supplier_id, description, total, paid_amount, money_kind, status, notes, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, 'CASH', %s, %s, %s, %s)
            """, (purchase_id, project_id, supplier_id,
                  f"Prepago material {supplier_name}",
                  info["prepago"], info["prepago"],
                  "DELIVERED" if total_delivered >= info["prepago"] else "PARTIAL",
                  f"Entregado: ${total_delivered:,.2f} / Falta: ${info['prepago'] - total_delivered:,.2f}",
                  now, now))

            # Create deliveries
            for wn, amount, del_date in deliveries:
                per_id = period_ids.get(wn)
                cur.execute("""
                    INSERT INTO material_deliveries (id, material_purchase_id, weekly_period_id, delivery_date, quantity, notes, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (cuid(), purchase_id, per_id, del_date, amount,
                      f"Entrega semana {wn}", now, now))

    print(f"Proveedores y prepagos insertados (El Gallo, Baez)")

    # --- 8. FIERRERO detail ---
    ws_fi = wb["FIERRERO"]
    fierrero_sections = [
        ("1ra Losa", 13, 18),
        ("2da Losa", 22, 27),
        ("3ra Losa", 31, 36),
    ]
    fierrero_count = 0
    # Fierrero items are not tied to a specific week - they're subcontract totals
    # We'll add them as notes on the project or as a reference
    # Actually, since the fierrero costs appear in the DESGLOZADO per-week,
    # let's add the detail breakdown as work items on a conceptual level
    # We'll attach them to the first week that has fierrero costs (week 8)
    first_fierrero_week = None
    for w in weeks_data:
        if w["fierrero"] > 0 and w["num"] in period_ids:
            first_fierrero_week = w["num"]
            break

    if first_fierrero_week:
        fi_period = period_ids[first_fierrero_week]
        for section_name, row_start, row_end in fierrero_sections:
            for r in range(row_start, row_end):
                desc = ws_fi.cell(r, 2).value
                unit = ws_fi.cell(r, 3).value
                volume = num(ws_fi.cell(r, 4).value)
                price = num(ws_fi.cell(r, 5).value)
                total = num(ws_fi.cell(r, 6).value)
                if not desc or total == 0:
                    continue
                cur.execute("""
                    INSERT INTO work_items (id, weekly_period_id, category, description, unit, volume, unit_price, total, money_kind, notes, created_at, updated_at)
                    VALUES (%s, %s, 'Fierrero', %s, %s, %s, %s, %s, 'CASH', %s, %s, %s)
                """, (cuid(), fi_period, f"{section_name} - {str(desc).strip()}",
                      str(unit).strip() if unit else "LOTE", volume, price, total,
                      "Desglose fierrero importado", now, now))
                fierrero_count += 1

        # Escalera
        esc_total = num(ws_fi.cell(39, 6).value)
        if esc_total > 0:
            cur.execute("""
                INSERT INTO work_items (id, weekly_period_id, category, description, unit, volume, unit_price, total, money_kind, notes, created_at, updated_at)
                VALUES (%s, %s, 'Fierrero', %s, 'PZA', 1, %s, %s, 'CASH', %s, %s, %s)
            """, (cuid(), fi_period, "Cimbra y armado de escalera",
                  esc_total, esc_total, "Desglose fierrero importado", now, now))
            fierrero_count += 1

    print(f"Insertados {fierrero_count} conceptos de fierrero")

    # --- Commit ---
    conn.commit()
    cur.close()
    conn.close()

    print("\n=== IMPORTACION COMPLETA ===")
    print(f"Proyecto: Real Alcazar 1370 ({built_m2} m2)")
    print(f"Semanas: {len(weeks_data)}")
    print(f"Destajos detallados (sem 27): {destajos_count}")
    print(f"Conceptos fierrero: {fierrero_count}")
    print(f"Proveedores: El Gallo, Baez")


if __name__ == "__main__":
    main()
